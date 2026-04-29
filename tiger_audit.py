"""TIGER/Line audit pipeline for SORTA MetroNow zones.

Read-only audit of OpenStreetMap ways with `tiger:reviewed=no` inside a
Hamilton County, OH MetroNow microtransit zone. Classifies defects and
emits XLSX, HTML dashboard, and CSV outputs.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import glob
import json
import math
import os
import sys
import time
from collections import defaultdict
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# All pipeline outputs are anchored to the script's directory, not CWD.
# Otherwise running `cd /elsewhere && python /path/to/tiger_audit.py` would
# scatter generated artifacts and Sheet 6's cross-zone status check would
# read a different tree than the one being written.
try:
    SCRIPT_DIR = Path(__file__).resolve().parent
except NameError:
    SCRIPT_DIR = Path.cwd()


ZONES = {
    "blue_ash_montgomery": {
        "name": "Blue Ash / Montgomery",
        "bbox": (39.16, -84.44, 39.24, -84.33),
        "description": "Blue Ash, Montgomery, Deer Park, Silverton, Kenwood, Madeira",
        "index_case_street": "O'Leary Avenue",
    },
    "springdale_sharonville": {
        "name": "Springdale / Sharonville",
        "bbox": (39.24, -84.48, 39.32, -84.38),
        "description": "Springdale, Sharonville, Glendale, Evendale, Lincoln Heights",
        "index_case_street": None,
    },
    "northgate_mt_healthy": {
        "name": "Northgate / Mt. Healthy",
        "bbox": (39.22, -84.58, 39.30, -84.48),
        "description": "Mt. Healthy, North College Hill, Finneytown, Northgate",
        "index_case_street": None,
    },
    "forest_park_pleasant_run": {
        "name": "Forest Park / Pleasant Run",
        "bbox": (39.26, -84.56, 39.34, -84.46),
        "description": "Forest Park, Pleasant Run, Greenhills",
        "index_case_street": None,
    },
}

OVERPASS_PRIMARY = "https://overpass-api.de/api/interpreter"
OVERPASS_MIRROR = "https://overpass.kumi.systems/api/interpreter"

CRITICAL = "CRITICAL"
HIGH = "HIGH"
LOW = "LOW"

CLASS_AB = "AB"
CLASS_A = "A"
CLASS_B = "B"
CLASS_C = "C"
CLASS_ORDER = [CLASS_AB, CLASS_A, CLASS_B, CLASS_C]


# ---------------------------------------------------------------------------
# Phase 1 - Fetch
# ---------------------------------------------------------------------------

def overpass_query(bbox: tuple[float, float, float, float]) -> str:
    s, w, n, e = bbox
    return (
        "[out:json][timeout:180];\n"
        f'way["highway"]["tiger:reviewed"="no"]\n'
        f"  ({s},{w},{n},{e});\n"
        "out tags geom;\n"
    )


def _utc_stamp() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")


OVERPASS_HEADERS = {
    "User-Agent": "tiger-audit-pipeline/1.0 (SORTA MetroNow OSM defect audit; contact: openstreetmap-volunteer)",
    "Accept": "application/json",
}


def _post_overpass(url: str, query: str) -> requests.Response:
    return requests.post(url, data={"data": query}, headers=OVERPASS_HEADERS, timeout=240)


CACHE_RETENTION_DAYS = 14


def _prune_old_cache(data_dir: Path, zone_key: str, keep_newest: int = 3) -> None:
    cached = sorted(
        data_dir.glob(f"{zone_key}_raw_*.json"),
        key=lambda p: p.stat().st_mtime,
    )
    if len(cached) <= keep_newest:
        return
    cutoff = time.time() - CACHE_RETENTION_DAYS * 86400
    for old in cached[:-keep_newest]:
        if old.stat().st_mtime < cutoff:
            try:
                old.unlink()
                print(f"  Pruned stale cache: {old.name}")
            except OSError as exc:
                print(f"  Could not prune {old.name}: {exc}")


SANITY_THRESHOLD = 100


def fetch_overpass(zone_key: str, out_dir: Path) -> dict:
    zone = ZONES[zone_key]
    query = overpass_query(zone["bbox"])
    data_dir = out_dir / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    last_error: Exception | None = None
    payload: dict | None = None

    attempts = [
        (OVERPASS_PRIMARY, 0),
        (OVERPASS_PRIMARY, 30),
        (OVERPASS_MIRROR, 0),
    ]

    for endpoint, presleep in attempts:
        if presleep:
            print(f"  Waiting {presleep}s before retry...")
            time.sleep(presleep)
        try:
            print(f"  POST {endpoint}")
            resp = _post_overpass(endpoint, query)
            if resp.status_code == 429:
                print("  HTTP 429 rate limit; sleeping 60s before next attempt")
                time.sleep(60)
                last_error = RuntimeError("429 rate limited")
                continue
            resp.raise_for_status()
            try:
                payload = resp.json()
            except ValueError as exc:
                snippet = resp.text[:500]
                raise RuntimeError(
                    f"Overpass response was not JSON. First 500 chars:\n{snippet}"
                ) from exc
            break
        except (
            requests.RequestException,
            ValueError,
            json.JSONDecodeError,
            RuntimeError,
        ) as exc:
            last_error = exc
            print(f"  Attempt failed: {exc}")
            continue

    fresh_fetch = payload is not None
    payload_source = "live Overpass response"
    if payload is None:
        # Sort by mtime, not filename — robust to timestamp format changes.
        cached = sorted(
            data_dir.glob(f"{zone_key}_raw_*.json"),
            key=lambda p: p.stat().st_mtime,
        )
        if cached:
            latest = cached[-1]
            age_s = time.time() - latest.stat().st_mtime
            age_label = f"{age_s/3600:.1f}h" if age_s < 86400 else f"{age_s/86400:.1f}d"
            print(
                f"Using cached data from {latest.name} (age {age_label}). "
                f"Live query failed."
            )
            if age_s > 14 * 86400:
                print(
                    f"  WARNING: cache is {age_label} old — re-run with network "
                    f"access for fresh data when possible."
                )
            with latest.open("r", encoding="utf-8") as fh:
                payload = json.load(fh)
            payload_source = f"cached file {latest}"
        else:
            raise RuntimeError(
                f"Overpass query failed and no cached data available: {last_error}"
            )

    # Validation + sanity check applies to BOTH fresh fetches and cached
    # payloads. A truncated cache file (manual edit, partial download, prior
    # under-threshold fetch that was saved without a tag, ...) needs to
    # surface the warning the same way a thin live response would. Recompute
    # the threshold flag here so it reflects this run's payload, not whatever
    # state was baked into the cache when it was written.
    if not isinstance(payload, dict):
        snippet = repr(payload)[:200]
        advice = (
            ""
            if fresh_fetch
            else " The cache may be corrupt or manually edited; delete the "
                 "file and re-run, or fix it."
        )
        raise RuntimeError(
            f"Overpass payload from {payload_source} is not a JSON object "
            f"(got {type(payload).__name__}).{advice} "
            f"First 200 chars of the payload: {snippet}"
        )
    payload.pop("_under_threshold", None)
    payload.pop("_element_count", None)
    elements = payload.get("elements")
    if not isinstance(elements, list):
        snippet = json.dumps(payload)[:500]
        raise RuntimeError(
            f"Overpass payload from {payload_source} is missing 'elements' "
            f"array (or it is not a list). First 500 chars:\n{snippet}"
        )
    if len(elements) < SANITY_THRESHOLD:
        print(
            f"  WARNING: only {len(elements)} elements (sanity threshold "
            f"{SANITY_THRESHOLD}) — audit may be based on truncated data."
        )
        payload["_under_threshold"] = True
        payload["_element_count"] = len(elements)

    if fresh_fetch:
        out_file = data_dir / f"{zone_key}_raw_{_utc_stamp()}.json"
        with out_file.open("w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=2)
        print(f"  Saved raw JSON to {out_file}")
        _prune_old_cache(data_dir, zone_key)

    n = len(payload.get("elements", []))
    print(f"Fetched {n} elements for {zone['name']}")
    return payload


# ---------------------------------------------------------------------------
# Phase 2 - Classify
# ---------------------------------------------------------------------------

def _norm_name(name: str | None) -> str | None:
    if not name:
        return None
    s = name.strip()
    if not s:
        return None
    return s.lower()


def _valid_latlon(lat: float, lon: float) -> bool:
    return -90.0 <= lat <= 90.0 and -180.0 <= lon <= 180.0


def _haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    if not (_valid_latlon(lat1, lon1) and _valid_latlon(lat2, lon2)):
        raise ValueError(
            f"Invalid lat/lon for haversine: ({lat1},{lon1}) ({lat2},{lon2})"
        )
    R = 6371000.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


GAP_THRESHOLD_M = 30.0
# A 3+ way junction where no node is shared produces a gap record per
# pair (3 ways -> 3 records). Cluster pairwise gap records that fall
# within this radius on the same street so each physical disconnect
# counts once. 5 m is wider than typical lane width but smaller than
# adjacent-junction spacing on residential streets.
GAP_CLUSTER_M = 5.0


def detect_gaps(class_b_streets: dict[str, list[dict]]) -> list[dict]:
    gaps: list[dict] = []
    for street, ways in class_b_streets.items():
        endpoints: list[tuple[dict, list[float], list[float]]] = []
        for w in ways:
            if len(w["geometry"]) >= 2:
                endpoints.append((w, w["geometry"][0], w["geometry"][-1]))

        # Phase 1: pairwise candidate gaps for every (way_i, way_j) on this street.
        raw_gaps_for_street: list[dict] = []
        for i in range(len(endpoints)):
            wi, si, ei = endpoints[i]
            for j in range(i + 1, len(endpoints)):
                wj, sj, ej = endpoints[j]
                pairs = [(si, sj), (si, ej), (ei, sj), (ei, ej)]
                best = None
                for (a, b) in pairs:
                    d = _haversine_m(a[0], a[1], b[0], b[1])
                    if best is None or d < best[0]:
                        best = (d, a, b)
                if best is None:
                    continue
                d, a, b = best
                if d <= 0.01:
                    continue  # already coincident -> no gap
                if d > GAP_THRESHOLD_M:
                    continue
                raw_gaps_for_street.append({
                    "lat": (a[0] + b[0]) / 2,
                    "lon": (a[1] + b[1]) / 2,
                    "street": street,
                    "type": "probable_disconnect",
                    "way1_id": wi["id"],
                    "way2_id": wj["id"],
                    "distance_m": round(d, 1),
                })

        # Phase 2: spatial cluster-by-junction. Sorted ascending by distance
        # so the tightest pair becomes each cluster's representative.
        raw_gaps_for_street.sort(key=lambda g: g["distance_m"])
        clustered_for_street: list[dict] = []
        for g in raw_gaps_for_street:
            collapsed = False
            for existing in clustered_for_street:
                if (
                    _haversine_m(g["lat"], g["lon"], existing["lat"], existing["lon"])
                    <= GAP_CLUSTER_M
                ):
                    collapsed = True
                    break
            if not collapsed:
                clustered_for_street.append(g)
        gaps.extend(clustered_for_street)
    return gaps


def classify(raw: dict) -> dict:
    elements = [e for e in raw.get("elements", []) if e.get("type") == "way"]
    by_norm: dict[str, list[dict]] = defaultdict(list)

    all_ways: list[dict] = []
    skipped_geom = 0
    for el in elements:
        tags = el.get("tags", {}) or {}
        name = tags.get("name")
        norm = _norm_name(name)
        geom = el.get("geometry") or []
        geom_pairs = [
            [g["lat"], g["lon"]]
            for g in geom
            if "lat" in g and "lon" in g and _valid_latlon(g["lat"], g["lon"])
        ]
        if not geom_pairs:
            skipped_geom += 1
        record = {
            "id": el.get("id"),
            "name": name,
            "name_display": name if name else "[Unnamed]",
            "name_key": norm,
            "highway": tags.get("highway"),
            "oneway": tags.get("oneway"),
            "tiger_reviewed": tags.get("tiger:reviewed"),
            "tiger_name_base": tags.get("tiger:name_base"),
            "tiger_cfcc": tags.get("tiger:cfcc"),
            "surface": tags.get("surface"),
            "lanes": tags.get("lanes"),
            "maxspeed": tags.get("maxspeed"),
            "geometry": geom_pairs,
        }
        all_ways.append(record)
        if norm is not None:
            by_norm[norm].append(record)

    class_b_norm_keys = {k for k, ways in by_norm.items() if len(ways) >= 2}

    for w in all_ways:
        is_a = w["highway"] == "residential" and w["oneway"] == "yes"
        is_b = w["name_key"] is not None and w["name_key"] in class_b_norm_keys
        if is_a and is_b:
            w["defect_class"] = CLASS_AB
            w["severity"] = CRITICAL
        elif is_a:
            w["defect_class"] = CLASS_A
            w["severity"] = CRITICAL
        elif is_b:
            w["defect_class"] = CLASS_B
            w["severity"] = HIGH
        else:
            w["defect_class"] = CLASS_C
            w["severity"] = LOW

    class_a = [w for w in all_ways if w["defect_class"] in (CLASS_A, CLASS_AB)]
    class_ab = [w for w in all_ways if w["defect_class"] == CLASS_AB]
    class_a_only = [w for w in all_ways if w["defect_class"] == CLASS_A]

    class_b_streets: dict[str, list[dict]] = {}
    for norm_key in class_b_norm_keys:
        ways = by_norm[norm_key]
        display = ways[0]["name"]
        class_b_streets[display] = ways

    by_highway: dict[str, int] = defaultdict(int)
    for w in all_ways:
        by_highway[w["highway"] or "(unset)"] += 1
    by_class: dict[str, int] = defaultdict(int)
    for w in all_ways:
        by_class[w["defect_class"]] += 1

    residential_count = sum(1 for w in all_ways if w["highway"] == "residential")
    oneway_yes_total = sum(1 for w in all_ways if w["oneway"] == "yes")

    gaps = detect_gaps(class_b_streets)

    if skipped_geom:
        print(
            f"  WARNING: {skipped_geom} ways had missing or invalid geometry "
            f"(rendered without polylines)."
        )

    summary_stats = {
        "total": len(all_ways),
        "residential": residential_count,
        "oneway_yes_total": oneway_yes_total,
        "class_a_count": len(class_a),
        "class_a_only_count": len(class_a_only),
        "class_ab_count": len(class_ab),
        "class_b_street_count": len(class_b_streets),
        "class_b_way_count": sum(
            1 for w in all_ways if w["defect_class"] in (CLASS_B, CLASS_AB)
        ),
        "gaps_found": len(gaps),
        "ways_missing_geom": skipped_geom,
        "under_sanity_threshold": bool(raw.get("_under_threshold", False)),
        "by_highway": dict(by_highway),
        "by_class": dict(by_class),
    }

    return {
        "all_ways": all_ways,
        "class_a": class_a,
        "class_a_only": class_a_only,
        "class_ab": class_ab,
        "class_b_streets": class_b_streets,
        "gaps": gaps,
        "summary_stats": summary_stats,
    }


# ---------------------------------------------------------------------------
# Phase 3 - XLSX
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
CRIT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
INDEX_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)
MONO_FONT = Font(name="Courier New", size=10)
CRIT_FONT = Font(name="Arial", size=10, color="9C0006")
HIGH_FONT = Font(name="Arial", size=10, color="9C5700")
OK_FONT = Font(name="Arial", size=10, color="006100")

THIN_BORDER = Border(bottom=Side(border_style="thin", color="B4C6E7"))
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)


def _style_header_row(ws, row_idx: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _style_data_cell(cell, font=DATA_FONT, align=DATA_ALIGN, border=THIN_BORDER) -> None:
    cell.font = font
    cell.alignment = align
    cell.border = border


def _autosize_columns(ws, max_widths: dict[int, int]) -> None:
    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 60)


def _severity_style(cell, severity: str) -> None:
    if severity == CRITICAL:
        cell.font = CRIT_FONT
        cell.fill = CRIT_FILL
    elif severity == HIGH:
        cell.font = HIGH_FONT
        cell.fill = HIGH_FILL
    elif severity == "OK":
        cell.font = OK_FONT
        cell.fill = OK_FILL


def _write_row(ws, row_idx: int, values: list, widths: dict[int, int]) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        _style_data_cell(cell)
        text_len = len(str(value)) if value is not None else 0
        widths[col_idx] = max(widths.get(col_idx, 0), text_len)


def write_xlsx(classified: dict, zone_key: str, out_path: Path, query_text: str, audit_ts: str) -> None:
    zone = ZONES[zone_key]
    stats = classified["summary_stats"]
    wb = Workbook()

    # ----- Sheet 1: Executive Summary -----
    ws = wb.active
    ws.title = "Executive Summary"
    ws.cell(row=1, column=1, value=f"TIGER/Line Audit - {zone['name']}").font = Font(
        name="Arial", size=14, bold=True, color="1F4E79"
    )
    ws.cell(row=1, column=2, value=f"Audit date (UTC): {audit_ts}").font = Font(
        name="Arial", size=10, italic=True
    )
    ws.cell(
        row=2,
        column=1,
        value=(
            f"Source: Overpass API live query | Bbox (S,W,N,E): {zone['bbox']}"
        ),
    ).font = Font(name="Arial", size=10, italic=True, color="555555")

    if stats.get("under_sanity_threshold"):
        warn = ws.cell(
            row=3,
            column=1,
            value=(
                f"WARNING: Overpass returned fewer than {SANITY_THRESHOLD} elements "
                f"for this zone. The audit may be based on a truncated or stale "
                f"dataset — re-run with network access before treating these "
                f"counts as authoritative."
            ),
        )
        warn.font = Font(name="Arial", size=10, bold=True, color="9C0006")
        warn.fill = CRIT_FILL
        warn.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)

    headers = ["Metric", "Count", "Significance"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    _style_header_row(ws, 4, len(headers))

    metrics = [
        ("Total unreviewed TIGER ways in zone", stats["total"],
         "All ways carrying tiger:reviewed=no in the bounding box."),
        ("Residential streets (unreviewed)", stats["residential"],
         "Subset where highway=residential."),
        ("Residential streets with oneway=yes (Class A)",
         sum(1 for w in classified["all_ways"] if w["highway"] == "residential" and w["oneway"] == "yes"),
         "Likely false one-way; routing engine sees no legal exit."),
        ("All road types with oneway=yes", stats["oneway_yes_total"],
         "Total one-way ways across all highway types."),
        ("Multi-segment disconnect risk streets (Class B)", stats["class_b_street_count"],
         "Named streets with 2+ unreviewed segments; intersection nodes may be disconnected."),
        ("Compound defect ways (Class AB)", stats["class_ab_count"],
         "False one-way AND on a multi-segment street; worst-case routing."),
    ]
    for i, (m, v, sig) in enumerate(metrics, start=5):
        ws.cell(row=i, column=1, value=m)
        c = ws.cell(row=i, column=2, value=v)
        c.number_format = "#,##0"
        ws.cell(row=i, column=3, value=sig)
        for col in (1, 2, 3):
            _style_data_cell(ws.cell(row=i, column=col))

    pct_row = 5 + len(metrics)
    ws.cell(row=pct_row, column=1, value="Residential % of total")
    pct = ws.cell(row=pct_row, column=2, value=f"=B6/B5")
    pct.number_format = "0.0%"
    ws.cell(row=pct_row, column=3,
            value="Share of unreviewed ways that are residential streets.")
    for col in (1, 2, 3):
        _style_data_cell(ws.cell(row=pct_row, column=col))

    ctx_start = pct_row + 2
    context_lines = [
        "CONTEXT",
        "SORTA operates MetroNow, a microtransit service in Hamilton County, Ohio,",
        "powered by Via Transportation. Via's routing engine consumes OpenStreetMap.",
        "The 2007-2008 TIGER/Line Census import seeded thousands of road segments",
        "with tiger:reviewed=no - meaning no human has verified them since import.",
        "False one-way tags and disconnected nodes inside this dataset translate",
        "directly into routing failures - service denials for transit-dependent riders.",
    ]
    for i, line in enumerate(context_lines):
        c = ws.cell(row=ctx_start + i, column=1, value=line)
        if i == 0:
            c.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
        else:
            c.font = Font(name="Arial", size=10, color="333333")

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A5"

    # ----- Sheet 2: Class A - Highest Risk (Class AB streets) -----
    ws2 = wb.create_sheet("Class A - Highest Risk")
    ab_by_street: dict[str, list[dict]] = defaultdict(list)
    for w in classified["class_ab"]:
        ab_by_street[w["name_display"]].append(w)
    ab_streets = sorted(ab_by_street.items(), key=lambda kv: kv[0].lower())
    headers2 = ["#", "Street", "Segments", "OSM Way IDs", "Severity", "Status"]
    for col, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=col, value=h)
    _style_header_row(ws2, 1, len(headers2))
    widths2: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers2)}
    for idx, (street, ways) in enumerate(ab_streets, start=1):
        row = idx + 1
        way_ids = ", ".join(str(w["id"]) for w in ways)
        values = [idx, street, len(ways), way_ids, CRITICAL, "Not started"]
        _write_row(ws2, row, values, widths2)
        _severity_style(ws2.cell(row=row, column=5), CRITICAL)
    ws2.freeze_panes = "A2"
    _autosize_columns(ws2, widths2)

    # ----- Sheet 3: Class A - Moderate Risk (Class A only) -----
    ws3 = wb.create_sheet("Class A - Moderate Risk")
    a_only_sorted = sorted(
        classified["class_a_only"], key=lambda w: (w["name_display"].lower(), w["id"] or 0)
    )
    headers3 = ["#", "Street", "OSM Way ID", "Severity", "Status"]
    for col, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=col, value=h)
    _style_header_row(ws3, 1, len(headers3))
    widths3: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers3)}
    index_street = (zone.get("index_case_street") or "").strip()
    for idx, w in enumerate(a_only_sorted, start=1):
        row = idx + 1
        is_index = bool(index_street) and (w["name"] or "").strip() == index_street
        status = "INDEX CASE" if is_index else "Not started"
        values = [idx, w["name_display"], w["id"], CRITICAL, status]
        _write_row(ws3, row, values, widths3)
        _severity_style(ws3.cell(row=row, column=4), CRITICAL)
        if is_index:
            for col in range(1, len(headers3) + 1):
                ws3.cell(row=row, column=col).fill = INDEX_FILL
    ws3.freeze_panes = "A2"
    _autosize_columns(ws3, widths3)

    # ----- Sheet 4: Class B - Multi-Segment -----
    ws4 = wb.create_sheet("Class B - Multi-Segment")
    headers4 = ["#", "Street", "Segments", "One-Way Segments", "Road Type", "Status"]
    for col, h in enumerate(headers4, start=1):
        ws4.cell(row=1, column=col, value=h)
    _style_header_row(ws4, 1, len(headers4))
    widths4: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers4)}
    big_b = [
        (street, ways)
        for street, ways in classified["class_b_streets"].items()
        if len(ways) >= 5
    ]
    big_b.sort(key=lambda kv: -len(kv[1]))
    for idx, (street, ways) in enumerate(big_b, start=1):
        row = idx + 1
        oneway_count = sum(1 for w in ways if w["oneway"] == "yes")
        types = sorted({w["highway"] or "(unset)" for w in ways})
        values = [idx, street, len(ways), oneway_count, ", ".join(types), "Not started"]
        _write_row(ws4, row, values, widths4)
        if oneway_count > 0:
            _severity_style(ws4.cell(row=row, column=4), CRITICAL)
    ws4.freeze_panes = "A2"
    _autosize_columns(ws4, widths4)

    # ----- Sheet 5: All Ways -----
    ws5 = wb.create_sheet("All Ways")
    headers5 = [
        "Way ID", "Street Name", "Highway Type", "Oneway",
        "Defect Class", "Severity", "tiger:cfcc", "tiger:name_base",
    ]
    for col, h in enumerate(headers5, start=1):
        ws5.cell(row=1, column=col, value=h)
    _style_header_row(ws5, 1, len(headers5))
    widths5: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers5)}
    class_rank = {c: i for i, c in enumerate(CLASS_ORDER)}
    sorted_ways = sorted(
        classified["all_ways"],
        key=lambda w: (class_rank[w["defect_class"]], w["name_display"].lower(), w["id"] or 0),
    )
    for idx, w in enumerate(sorted_ways, start=2):
        values = [
            w["id"], w["name_display"], w["highway"] or "", w["oneway"] or "",
            w["defect_class"], w["severity"],
            w["tiger_cfcc"] or "", w["tiger_name_base"] or "",
        ]
        _write_row(ws5, idx, values, widths5)
        _severity_style(ws5.cell(row=idx, column=6), w["severity"])
    ws5.freeze_panes = "A2"
    _autosize_columns(ws5, widths5)

    # ----- Sheet 6: MetroNow Zones -----
    # Cross-zone awareness: a workbook for any one zone marks the OTHER
    # three zones as AUDIT COMPLETE if their output CSVs exist on disk
    # (an existing all_ways.csv is the cheapest, most reliable signal that
    # the zone has been audited at some point).
    ws6 = wb.create_sheet("MetroNow Zones")
    headers6 = ["Zone", "Audit Status", "Unreviewed Segments", "Notes"]
    for col, h in enumerate(headers6, start=1):
        ws6.cell(row=1, column=col, value=h)
    _style_header_row(ws6, 1, len(headers6))
    widths6: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers6)}
    for idx, (k, z) in enumerate(ZONES.items(), start=2):
        if k == zone_key:
            status = "AUDIT COMPLETE"
            seg: int | str = stats["total"]
            severity_for_row = "OK"
        else:
            # Read from the same SCRIPT_DIR-anchored tree the pipeline writes to.
            other_csv = SCRIPT_DIR / f"tiger_audit_{k}" / "csv" / "all_ways.csv"
            if other_csv.exists():
                try:
                    with other_csv.open("r", encoding="utf-8") as fh:
                        # Subtract one for the header row.
                        other_count = max(sum(1 for _ in fh) - 1, 0)
                    status = "AUDIT COMPLETE"
                    seg = other_count
                    severity_for_row = "OK"
                except OSError:
                    status = "NOT STARTED"
                    seg = ""
                    severity_for_row = HIGH
            else:
                status = "NOT STARTED"
                seg = ""
                severity_for_row = HIGH
        values = [z["name"], status, seg, z["description"]]
        _write_row(ws6, idx, values, widths6)
        _severity_style(ws6.cell(row=idx, column=2), severity_for_row)
    ws6.freeze_panes = "A2"
    _autosize_columns(ws6, widths6)

    # ----- Sheet 7: Work Plan -----
    ws7 = wb.create_sheet("Work Plan")
    headers7 = [
        "Phase", "Description", "Items", "Min/Item", "Hours",
    ]
    for col, h in enumerate(headers7, start=1):
        ws7.cell(row=1, column=col, value=h)
    _style_header_row(ws7, 1, len(headers7))
    widths7: dict[int, int] = {i + 1: len(h) for i, h in enumerate(headers7)}
    p4_items = max(stats["residential"] - stats["class_a_count"], 0)
    work_rows = [
        ("P1", "Fix compound defects (Class AB)", stats["class_ab_count"], 12),
        ("P2", "Fix single-segment false one-ways (Class A only)", stats["class_a_only_count"], 4),
        ("P3", "Inspect multi-segment disconnects (Class B)",
         stats["class_b_street_count"], 7),
        ("P4", "Review remaining residential ways", p4_items, 2),
    ]
    for i, (phase, desc, items, mins) in enumerate(work_rows, start=2):
        ws7.cell(row=i, column=1, value=phase)
        ws7.cell(row=i, column=2, value=desc)
        ws7.cell(row=i, column=3, value=items).number_format = "#,##0"
        ws7.cell(row=i, column=4, value=mins).number_format = "#,##0"
        formula = f"=C{i}*D{i}/60"
        h_cell = ws7.cell(row=i, column=5, value=formula)
        h_cell.number_format = "#,##0.0"
        for col in range(1, len(headers7) + 1):
            _style_data_cell(ws7.cell(row=i, column=col))
            widths7[col] = max(widths7.get(col, 0), len(str(ws7.cell(row=i, column=col).value)))
    total_row = len(work_rows) + 2
    ws7.cell(row=total_row, column=1, value="Total")
    ws7.cell(row=total_row, column=1).font = Font(name="Arial", size=10, bold=True)
    ws7.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row - 1})").number_format = "#,##0.0"
    for col in (1, 5):
        _style_data_cell(ws7.cell(row=total_row, column=col))
    ws7.freeze_panes = "A2"
    _autosize_columns(ws7, widths7)

    # ----- Sheet 8: Overpass Query -----
    ws8 = wb.create_sheet("Overpass Query")
    info_lines = [
        ("Endpoint", OVERPASS_PRIMARY),
        ("Bounding box (S, W, N, E)", str(zone["bbox"])),
        ("Executed (UTC)", audit_ts),
        ("", ""),
        ("Query", ""),
    ]
    for i, (k, v) in enumerate(info_lines, start=1):
        kc = ws8.cell(row=i, column=1, value=k)
        vc = ws8.cell(row=i, column=2, value=v)
        kc.font = Font(name="Arial", size=10, bold=True)
        vc.font = MONO_FONT
    base = len(info_lines) + 1
    for i, line in enumerate(query_text.splitlines()):
        c = ws8.cell(row=base + i, column=1, value=line)
        c.font = MONO_FONT
    ws8.column_dimensions["A"].width = 60
    ws8.column_dimensions["B"].width = 60

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  XLSX saved: {out_path}")


# ---------------------------------------------------------------------------
# Phase 4 - HTML Dashboard
# ---------------------------------------------------------------------------

DASHBOARD_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>TIGER Audit Dashboard - __ZONE_NAME__</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.heat@0.2.0/dist/leaflet-heat.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<style>
:root{
  --bg:#f8f9fa;--sidebar-bg:#fff;--text:#333;--card-bg:#f8f9fa;
  --header-bg:#1F4E79;--header-text:#fff;--border:#ddd;--border-soft:#eee;
  --accent:#1F4E79;--link:#0563C1;--muted:#666;--input-bg:#fff;
}
:root.dark{
  --bg:#1a1a2e;--sidebar-bg:#16213e;--text:#e0e0e0;--card-bg:#0f3460;
  --header-bg:#0a1929;--header-text:#e0e0e0;--border:#2a3a5a;--border-soft:#1f2e4d;
  --accent:#4a7fb8;--link:#80bfff;--muted:#a0a0c0;--input-bg:#0f3460;
}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:Arial,sans-serif;background:var(--bg);color:var(--text)}
#header{background:var(--header-bg);color:var(--header-text);padding:10px 16px;display:flex;justify-content:space-between;align-items:center;z-index:1000;height:42px}
#header h1{font-size:16px}
#header .sub{font-size:12px;opacity:.85}
#main{display:flex;height:calc(100vh - 42px);position:relative}
#sidebar{width:300px;background:var(--sidebar-bg);overflow-y:auto;border-right:1px solid var(--border);flex-shrink:0;font-size:13px;color:var(--text);transition:width .25s ease,min-width .25s ease}
#sidebar.collapsed{width:0;min-width:0;border-right:none;overflow:hidden}
#sbToggle{position:absolute;top:50%;left:300px;transform:translate(-50%,-50%);width:22px;height:60px;background:var(--accent);color:#fff;border:none;border-radius:0 6px 6px 0;cursor:pointer;z-index:1000;font-size:13px;line-height:1;box-shadow:2px 0 4px rgba(0,0,0,0.15);transition:left .25s ease;padding:0;display:flex;align-items:center;justify-content:center}
#sbToggle:hover{filter:brightness(1.15)}
#sidebar.collapsed ~ #sbToggle{left:0;border-radius:0 6px 6px 0;transform:translate(0,-50%)}
#sidebar .sec{padding:12px 14px;border-bottom:1px solid var(--border-soft)}
#sidebar .sec h3{font-size:13px;color:var(--accent);margin-bottom:8px;text-transform:uppercase;letter-spacing:.5px;display:flex;align-items:center;justify-content:space-between}
.kg{display:grid;grid-template-columns:1fr 1fr;gap:8px}
.kp{background:var(--card-bg);border-radius:6px;padding:8px 10px;border-left:3px solid var(--accent)}
.kp.cr{border-left-color:#C00000}.kp.hi{border-left-color:#ED7D31}.kp.nw{border-left-color:#548235}
.kp .n{font-size:20px;font-weight:bold;color:var(--accent)}.kp .l{font-size:10px;color:var(--muted);text-transform:uppercase}
.lt{display:flex;align-items:center;gap:6px;padding:4px 0;cursor:pointer}
.lt input{cursor:pointer}.lt label{cursor:pointer;font-size:12px;color:var(--text)}
.sw{display:inline-block;border-radius:2px}
#search{width:100%;padding:6px 8px;border:1px solid var(--border);border-radius:4px;font-size:12px;background:var(--input-bg);color:var(--text)}
#search:focus{outline:none;border-color:var(--accent)}
.eb{display:inline-block;padding:5px 10px;background:var(--accent);color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:11px;margin:2px}
.eb:hover{filter:brightness(1.15)}
#vs{font-size:11px;color:var(--muted);line-height:1.6}
#vs b{color:var(--text)}
#map{flex:1;background:var(--bg)}
.leaflet-popup-content{max-width:340px;font-size:12px;line-height:1.5;color:#333}
.leaflet-popup-content a{color:#0563C1}
.tg{display:inline-block;padding:1px 6px;border-radius:3px;font-size:10px;font-weight:bold;margin:2px 0}
.tg.ab{background:#FFC7CE;color:#9C0006}.tg.a{background:#FFC7CE;color:#9C0006}
.tg.b{background:#FFEB9C;color:#9C5700}.tg.c{background:#E8E8E8;color:#666}
select#bmsel,select#waybackRel{background:var(--input-bg);color:var(--text);border:1px solid var(--border)}

/* R3 loading overlay */
#loading{position:fixed;inset:0;background:rgba(255,255,255,0.95);z-index:9999;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:16px;color:#1F4E79;font-weight:bold;font-family:Arial,sans-serif}
.spinner{width:48px;height:48px;border:4px solid #e0e0e0;border-top-color:#1F4E79;border-radius:50%;animation:spin 1s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}

/* R2 cluster icon */
.gcl{background:#9C27B0;color:#fff;border-radius:50%;width:36px;height:36px;display:flex;align-items:center;justify-content:center;font-weight:bold;font-size:12px;border:2px solid #fff;box-shadow:0 1px 4px rgba(0,0,0,0.3)}
.marker-cluster-small,.marker-cluster-medium,.marker-cluster-large{background:none}
.marker-cluster-small div,.marker-cluster-medium div,.marker-cluster-large div{background:none}

/* R4 mobile sidebar */
#hamburger{display:none;position:fixed;top:50px;left:12px;z-index:1001;background:var(--accent);color:#fff;border:none;padding:8px 12px;border-radius:4px;cursor:pointer;font-size:16px;box-shadow:0 2px 8px rgba(0,0,0,0.3)}
#backdrop{display:none;position:fixed;top:42px;left:0;right:0;bottom:0;background:rgba(0,0,0,0.4);z-index:999}
#backdrop[hidden]{display:none}
@media (max-width:768px){
  #sidebar{position:fixed;left:-300px;top:42px;height:calc(100vh - 42px);z-index:1000;transition:left 0.3s ease;box-shadow:2px 0 8px rgba(0,0,0,0.3);width:300px !important}
  #sidebar.open{left:0}
  #sbToggle{display:none}
  #hamburger{display:block}
  #backdrop:not([hidden]){display:block}
  #main{display:block}
  #map{height:calc(100vh - 42px);width:100%}
  /* Hide zoom control on small screens — it overlaps the hamburger and is rarely needed when pinch-zoom is available */
  .leaflet-control-zoom{display:none}
}
/* Visible focus indicator for keyboard navigation */
.leaflet-interactive:focus,#sr [role="option"]:focus{outline:3px solid #FFB800;outline-offset:2px}
#sr [role="option"][aria-selected="true"]{background:rgba(255,184,0,0.18);outline:1px dashed var(--accent)}

/* E4 pulse */
@keyframes pulseStroke{0%,100%{stroke-opacity:1}50%{stroke-opacity:0.4}}
.pulse{animation:pulseStroke 1s ease-in-out infinite}
@media (prefers-reduced-motion:reduce){.spinner{animation:none}.pulse{animation:none;stroke-opacity:1}}

/* E2 help overlay */
#help{position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);background:var(--sidebar-bg);color:var(--text);border:1px solid var(--border);border-radius:8px;padding:20px 24px;z-index:2000;box-shadow:0 4px 24px rgba(0,0,0,0.4);max-width:380px;font-size:13px}
#help[hidden]{display:none}
#help h3{color:var(--accent);margin-bottom:12px;font-size:14px}
#help table{width:100%;border-collapse:collapse;margin-bottom:8px}
#help td{padding:4px 8px;vertical-align:top}
#help td:first-child{color:var(--accent);font-family:monospace;white-space:nowrap;width:110px}
#help .hint{font-size:11px;color:var(--muted);margin-top:8px;text-align:center}

/* E3 status toast */
#status{position:fixed;bottom:24px;left:50%;transform:translateX(-50%);background:var(--header-bg);color:var(--header-text);padding:10px 18px;border-radius:6px;z-index:2000;box-shadow:0 4px 12px rgba(0,0,0,0.3);font-size:13px;max-width:80vw;text-align:center}
#status[hidden]{display:none}

/* E5 print mode */
#printBar{display:none;position:fixed;top:0;left:0;right:0;background:#1F4E79;color:#fff;padding:12px;z-index:2000;text-align:center;font-size:13px}
.printing #sidebar,.printing #header,.printing #hamburger,.printing #sbToggle,.printing .leaflet-control-zoom,.printing .leaflet-control-attribution{display:none !important}
.printing #map{width:100vw;height:100vh}
.printing #printBar{display:block}
@media print{
  :root{--bg:#fff;--sidebar-bg:#fff;--text:#000;--card-bg:#fff;--header-bg:#1F4E79;--border:#ccc;--accent:#1F4E79;--header-text:#fff}
  body{-webkit-print-color-adjust:exact;print-color-adjust:exact}
  #sidebar,#header,#hamburger,#sbToggle,.leaflet-control-zoom,.leaflet-control-attribution,#status,#help{display:none !important}
  #map{width:100vw;height:100vh}
  #printBar{display:block !important}
}

/* E6 dark toggle */
#dmt{background:none;border:1px solid var(--border);color:var(--accent);cursor:pointer;font-size:14px;padding:2px 8px;border-radius:4px;line-height:1.4}
#dmt:hover{filter:brightness(1.2)}

/* R1 street popup */
.pp-h{font-weight:bold;color:#1F4E79;margin-bottom:6px;font-size:13px}
.pp-bd{font-size:11px;color:#666;margin-bottom:6px}
.pp-clicked{background:#FFF8DC;padding:4px 6px;border-radius:3px;font-size:11px;margin-bottom:6px;border-left:3px solid #ED7D31}
.pp-list{max-height:180px;overflow-y:auto;border-top:1px solid #eee;padding-top:4px;margin-top:4px}
.pp-row{display:flex;align-items:center;gap:6px;padding:3px 0;font-size:11px;border-bottom:1px solid #f5f5f5}
.pp-foot{margin-top:8px;display:flex;gap:8px;align-items:center;font-size:11px;flex-wrap:wrap}
.pp-foot a{cursor:pointer}
.pp-btn{display:inline-block;padding:3px 8px;background:#1F4E79;color:#fff;border:none;border-radius:3px;cursor:pointer;font-size:11px;text-decoration:none}
.pp-btn:hover{filter:brightness(1.15)}
</style>
</head>
<body>
<div id="loading" role="status" aria-live="polite">
  <div class="spinner" aria-hidden="true"></div>
  <div>Rendering __TOTAL__ segments&hellip;</div>
</div>
<div id="header">
<div><h1>TIGER Audit Dashboard &mdash; __ZONE_NAME__</h1>
<div class="sub">__TOTAL__ unreviewed road segments &middot; Full geometry + node gap analysis &middot; __AUDIT_TS__</div></div>
</div>
<button id="hamburger" aria-label="Toggle sidebar" aria-controls="sidebar" aria-expanded="false">&#9776;</button>
<div id="backdrop" role="presentation" hidden></div>
<div id="main">
<div id="sidebar" role="region" aria-label="Dashboard controls">
<div class="sec"><h3>Zone Totals <button id="dmt" aria-label="Switch to dark mode" title="Switch to dark mode">&#127769;</button></h3>
<div class="kg">
<div class="kp"><div class="n" id="kt" aria-live="polite">-</div><div class="l">Total</div></div>
<div class="kp"><div class="n" id="kr" aria-live="polite">-</div><div class="l">Residential</div></div>
<div class="kp cr"><div class="n" id="kab" aria-live="polite">-</div><div class="l">Class AB</div></div>
<div class="kp cr"><div class="n" id="ka" aria-live="polite">-</div><div class="l">Class A</div></div>
<div class="kp hi"><div class="n" id="kb" aria-live="polite">-</div><div class="l">Class B</div></div>
<div class="kp nw"><div class="n" id="kg" aria-live="polite">-</div><div class="l">Node Gaps</div></div>
</div></div>
<div class="sec"><h3>Layers</h3>
<div class="lt"><input type="checkbox" id="lab" checked aria-label="Toggle Class AB Compound layer"><span class="sw" style="background:#C00000;width:14px;height:4px" aria-hidden="true"></span><label for="lab">Class AB Compound (critical)</label></div>
<div class="lt"><input type="checkbox" id="la" checked aria-label="Toggle Class A False one-way layer"><span class="sw" style="background:#FF4444;width:14px;height:4px" aria-hidden="true"></span><label for="la">Class A False one-way</label></div>
<div class="lt"><input type="checkbox" id="lb" checked aria-label="Toggle Class B Multi-segment layer"><span class="sw" style="background:#ED7D31;width:14px;height:4px" aria-hidden="true"></span><label for="lb">Class B Multi-segment</label></div>
<div class="lt"><input type="checkbox" id="lc" aria-label="Toggle Class C Unreviewed layer"><span class="sw" style="background:#999;width:14px;height:4px" aria-hidden="true"></span><label for="lc">Class C Unreviewed</label></div>
<div class="lt"><input type="checkbox" id="lgap" checked aria-label="Toggle Node disconnect points layer"><span class="sw" style="background:#9C27B0;width:10px;height:10px;border-radius:50%" aria-hidden="true"></span><label for="lgap">Node disconnect points</label></div>
<div class="lt"><input type="checkbox" id="lheat" aria-label="Toggle Density heatmap layer"><span class="sw" style="background:linear-gradient(90deg,blue,yellow,red);width:30px;height:8px" aria-hidden="true"></span><label for="lheat">Density heatmap</label></div>
</div>
<div class="sec"><h3>Basemap</h3>
<select id="bmsel" aria-label="Select basemap" style="width:100%;padding:6px;border-radius:4px;font-size:12px">
<optgroup label="Vector / road">
<option value="carto-voyager">CARTO Voyager (default)</option>
<option value="carto-positron">CARTO Positron (light)</option>
<option value="carto-darkmatter">CARTO Dark Matter</option>
<option value="osm">OpenStreetMap Standard</option>
</optgroup>
<optgroup label="Imagery (free)">
<option value="esri-imagery">Esri World Imagery</option>
<option value="esri-clarity">Esri World Imagery (Clarity)</option>
<option value="esri-wayback">Esri Wayback (date selector)</option>
<option value="usgs-imagery">USGS National Map - Imagery</option>
<option value="osip">Ohio OSIP (6-inch, best-effort)</option>
</optgroup>
<optgroup label="Topo">
<option value="usgs-topo">USGS National Map - Topo</option>
<option value="esri-topo">Esri World Topographic</option>
</optgroup>
</select>
</div>
<div class="sec" id="waybackSec" style="display:none">
<h3>Wayback Release</h3>
<select id="waybackRel" aria-label="Select Wayback release date" style="width:100%;padding:6px;border-radius:4px;font-size:12px">
<option value="">Loading available releases...</option>
</select>
<div style="font-size:10px;color:var(--muted);margin-top:4px;line-height:1.4">Each entry is a dated snapshot of Esri World Imagery. Useful for confirming whether a TIGER defect was already wrong on recent imagery vs. earlier captures. <a href="https://livingatlas.arcgis.com/wayback/" target="_blank" rel="noopener">About Wayback</a></div>
</div>
<div class="sec"><h3>Search</h3>
<input type="text" id="search" placeholder="Street name..." aria-label="Search streets" role="combobox" aria-expanded="false" aria-controls="sr" aria-autocomplete="list">
<div id="sr" role="listbox" aria-label="Search results" style="margin-top:6px;max-height:120px;overflow-y:auto"></div>
</div>
<div class="sec"><h3>Viewport Stats</h3><div id="vs" aria-live="polite">Pan/zoom to update</div></div>
<div class="sec"><h3>Export Visible</h3>
<button class="eb" onclick="xGeo()" aria-label="Export visible segments as GeoJSON">GeoJSON</button>
<button class="eb" onclick="xCSV()" aria-label="Export visible segments as CSV">CSV</button>
<button class="eb" onclick="loadInJOSM()" aria-label="Load all visible ways into JOSM" title="Load visible ways into JOSM via Remote Control">Load in JOSM</button>
<button class="eb" onclick="printView()" aria-label="Open print view" title="Open print view">Print View</button>
</div>
<div class="sec" style="font-size:10px;color:var(--muted)">Press <b>?</b> for keyboard shortcuts</div>
</div>
<button id="sbToggle" type="button" title="Collapse sidebar" aria-label="Toggle sidebar">&#9664;</button>
<div id="map" role="region" aria-label="Map of TIGER audit segments"></div>
</div>
<div id="status" hidden role="status" aria-live="polite"></div>
<div id="help" hidden role="dialog" aria-label="Keyboard shortcuts">
<h3>Keyboard Shortcuts</h3>
<table>
<tr><td>/ or Ctrl+K</td><td>Focus search</td></tr>
<tr><td>Esc</td><td>Clear search / highlight / popup</td></tr>
<tr><td>1&ndash;6</td><td>Toggle layers (AB, A, B, C, Gaps, Heatmap)</td></tr>
<tr><td>S</td><td>Toggle satellite (Esri Imagery)</td></tr>
<tr><td>?</td><td>Show / hide this help</td></tr>
</table>
<div class="hint">Press <b>?</b> or <b>Esc</b> to close</div>
</div>
<div id="printBar"></div>
<script>
var D=__DATA_JSON__;
var CL={AB:'#C00000',A:'#FF4444',B:'#ED7D31',C:'#999999'};
var CW={AB:4,A:3,B:2.5,C:1.5};
var CN={AB:'Compound (A+B)',A:'False One-Way',B:'Multi-Segment',C:'Unreviewed'};
var LK={lab:'AB',la:'A',lb:'B',lc:'C',lgap:'gap',lheat:'heat'};
var LR={AB:'lab',A:'la',B:'lb',C:'lc',gap:'lgap',heat:'lheat'};
var ZONE_NAME='__ZONE_NAME__';

var map=L.map('map',{zoomControl:true}).setView(__CENTER__,13);

var BMS={
  'carto-voyager':{url:'https://{s}.basemaps.cartocdn.com/rastertiles/voyager/{z}/{x}/{y}{r}.png',opts:{subdomains:'abcd',maxZoom:20,attribution:'&copy; OSM, &copy; CARTO'}},
  'carto-positron':{url:'https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png',opts:{subdomains:'abcd',maxZoom:20,attribution:'&copy; OSM, &copy; CARTO'}},
  'carto-darkmatter':{url:'https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png',opts:{subdomains:'abcd',maxZoom:20,attribution:'&copy; OSM, &copy; CARTO'}},
  'osm':{url:'https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',opts:{subdomains:'abc',maxZoom:19,attribution:'&copy; OpenStreetMap contributors'}},
  'esri-imagery':{url:'https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',opts:{maxZoom:19,attribution:'Tiles &copy; Esri &mdash; Source: Esri, Maxar, Earthstar Geographics, USDA, USGS, AeroGRID, IGN'}},
  'esri-clarity':{url:'https://clarity.maptiles.arcgis.com/arcgis/rest/services/World_Imagery_Firefly/MapServer/tile/{z}/{y}/{x}',opts:{maxZoom:19,attribution:'Esri Clarity / Firefly'}},
  'usgs-imagery':{url:'https://basemap.nationalmap.gov/arcgis/rest/services/USGSImageryOnly/MapServer/tile/{z}/{y}/{x}',opts:{maxZoom:18,attribution:'USGS National Map: Imagery'}},
  'usgs-topo':{url:'https://basemap.nationalmap.gov/arcgis/rest/services/USGSTopo/MapServer/tile/{z}/{y}/{x}',opts:{maxZoom:18,attribution:'USGS National Map: Topo'}},
  'esri-topo':{url:'https://server.arcgisonline.com/ArcGIS/rest/services/World_Topo_Map/MapServer/tile/{z}/{y}/{x}',opts:{maxZoom:19,attribution:'Tiles &copy; Esri'}},
  'osip':{url:'https://gis5.oit.ohio.gov/arcgis/rest/services/OSIP/OSIP_Latest/ImageServer/tile/{z}/{y}/{x}',opts:{maxZoom:20,attribution:'OGRIP / Ohio Statewide Imagery Program'}}
};

var curBM=null,curBMKey='carto-voyager',prevBMKey=null;
var WAYBACK_CONFIG_URL='https://s3-us-west-2.amazonaws.com/config.maptiles.arcgis.com/waybackconfig.json';
var waybackConfig=null,waybackLoading=null;

function loadWaybackConfig(){
  if(waybackConfig)return Promise.resolve(waybackConfig);
  if(waybackLoading)return waybackLoading;
  waybackLoading=fetch(WAYBACK_CONFIG_URL,{cache:'force-cache'})
    .then(function(r){if(!r.ok)throw new Error('HTTP '+r.status);return r.json();})
    .then(function(j){waybackConfig=j;return j;})
    .catch(function(err){console.warn('Wayback config fetch failed:',err);waybackConfig={};return {};});
  return waybackLoading;
}

function populateWaybackSelect(){
  var sel=document.getElementById('waybackRel');
  loadWaybackConfig().then(function(cfg){
    var keys=Object.keys(cfg);
    if(!keys.length){
      sel.innerHTML='<option value="">Could not load releases &mdash; try a Wayback release ID from livingatlas.arcgis.com/wayback</option>';
      return;
    }
    var entries=keys.map(function(k){
      var rec=cfg[k]||{};
      var date=rec.itemReleaseName||rec.releaseDateLabel||rec.itemTitle||k;
      return {rel:k,date:String(date)};
    }).sort(function(a,b){return a.date<b.date?1:(a.date>b.date?-1:0);});
    var saved='';try{saved=localStorage.getItem('tigerWaybackRel')||'';}catch(e){}
    sel.innerHTML='';
    entries.forEach(function(e){
      var o=document.createElement('option');
      o.value=e.rel;o.textContent=e.date;sel.appendChild(o);
    });
    if(saved&&sel.querySelector('option[value="'+saved+'"]'))sel.value=saved;
    else sel.value=entries[0].rel;
    if(document.getElementById('bmsel').value==='esri-wayback')setBasemap('esri-wayback');
  });
}

function updateWaybackVis(){
  var key=document.getElementById('bmsel').value;
  document.getElementById('waybackSec').style.display=(key==='esri-wayback')?'':'none';
  if(key==='esri-wayback'&&!waybackConfig)populateWaybackSelect();
}

function setBasemap(key){
  if(curBM){map.removeLayer(curBM);curBM=null;}
  curBMKey=key;
  updateWaybackVis();
  var url, opts;
  if(key==='esri-wayback'){
    var rel=document.getElementById('waybackRel').value;
    if(!rel){
      var b0=BMS['esri-imagery'];url=b0.url;opts=b0.opts;
    } else {
      url='https://wayback.maptiles.arcgis.com/arcgis/rest/services/World_Imagery/WMTS/1.0.0/default028mm/MapServer/tile/'+encodeURIComponent(rel)+'/{z}/{y}/{x}';
      opts={maxZoom:19,attribution:'Esri Wayback Imagery (release '+rel+')'};
      try{localStorage.setItem('tigerWaybackRel',rel);}catch(e){}
    }
  } else {
    var b=BMS[key]||BMS['carto-voyager'];
    url=b.url;opts=b.opts;
  }
  curBM=L.tileLayer(url,opts).addTo(map);
  curBM.bringToBack();
  try{localStorage.setItem('tigerBmKey',key);}catch(e){}
  scheduleHashWrite();
}

var ly={AB:L.layerGroup(),A:L.layerGroup(),B:L.layerGroup(),C:L.layerGroup()};
var gL=L.markerClusterGroup({
  disableClusteringAtZoom:16,maxClusterRadius:50,
  iconCreateFunction:function(cluster){
    var n=cluster.getChildCount();
    return L.divIcon({html:'<div class="gcl" role="button" aria-label="'+n+' gap markers, click to expand">'+n+'</div>',className:'',iconSize:L.point(36,36)});
  }
});
gL.on('clusterclick',function(e){if(e.originalEvent)L.DomEvent.stopPropagation(e.originalEvent);});
var hL=null,aP=[];
var streetIndex={},wayIndex={};
var hl={type:null,key:null};
var gapPairLine=null;
var restoring=false,hashTimer;
var helpOpen=false;
var josmProbeOk=null;

document.getElementById('kt').textContent=D.stats.total.toLocaleString();
document.getElementById('kr').textContent=D.stats.residential.toLocaleString();
document.getElementById('kab').textContent=D.stats.class_ab.toLocaleString();
document.getElementById('ka').textContent=D.stats.class_a.toLocaleString();
document.getElementById('kb').textContent=D.stats.class_b.toLocaleString();
document.getElementById('kg').textContent=D.stats.gaps_found.toLocaleString();

function escAttr(s){return String(s).replace(/&/g,'&amp;').replace(/"/g,'&quot;').replace(/</g,'&lt;');}

function _resetStyles(){
  aP.forEach(function(p){
    p.setStyle({weight:CW[p.wd.c],opacity:p.wd.c==='C'?0.4:0.8});
    var el=p.getElement();if(el)el.classList.remove('pulse');
  });
  if(gapPairLine){map.removeLayer(gapPairLine);gapPairLine=null;}
}
function clearHighlight(){_resetStyles();hl={type:null,key:null};scheduleHashWrite();}
function findNearMissEndpoints(g1,g2){
  var s1=g1[0],e1=g1[g1.length-1],s2=g2[0],e2=g2[g2.length-1];
  var pairs=[[s1,s2],[s1,e2],[e1,s2],[e1,e2]];
  var min=Infinity,best=null;
  pairs.forEach(function(pr){var dx=pr[0][0]-pr[1][0],dy=pr[0][1]-pr[1][1],d=dx*dx+dy*dy;if(d<min){min=d;best=pr;}});
  return best;
}
function setHighlight(type,key){
  _resetStyles();
  hl={type:type,key:key};
  if(type==='street'){
    var name=key;
    aP.forEach(function(p){
      if(p.wd.n===name){p.setStyle({weight:CW[p.wd.c]+2,opacity:1.0});p.bringToFront();}
      else p.setStyle({opacity:0.2});
    });
  }else if(type==='gap'){
    var p1=wayIndex[key[0]],p2=wayIndex[key[1]];
    if(p1&&p2){
      [p1,p2].forEach(function(p){
        p.setStyle({weight:CW[p.wd.c]+3,opacity:1.0});
        p.bringToFront();
        var el=p.getElement();if(el)el.classList.add('pulse');
      });
      var ep=findNearMissEndpoints(p1.wd.g,p2.wd.g);
      gapPairLine=L.polyline(ep,{color:'#FF00FF',weight:3,dashArray:'5,10',interactive:false}).addTo(map);
      map.fitBounds(L.latLngBounds(p1.wd.g.concat(p2.wd.g)),{padding:[40,40],maxZoom:18});
    }
  }
  scheduleHashWrite();
}

function openStreetPopup(w,latlng){
  var name=w.n||'[Unnamed]';
  var segs=w.n?(streetIndex[name]||[]):[wayIndex[w.id]];
  var bd={AB:0,A:0,B:0,C:0};
  segs.forEach(function(p){bd[p.wd.c]++;});
  var bdStr=Object.keys(bd).filter(function(k){return bd[k]>0;}).map(function(k){return bd[k]+' '+k;}).join(' &middot; ');
  var rows=segs.map(function(p){
    var ww=p.wd;
    var idLnk='<a href="https://www.openstreetmap.org/way/'+ww.id+'" target="_blank" rel="noopener">'+ww.id+'</a>';
    var iD='<a href="https://www.openstreetmap.org/edit?editor=id#map=18/'+ww.g[0][0]+'/'+ww.g[0][1]+'" target="_blank" rel="noopener">iD</a>';
    var jo='<a href="http://localhost:8111/load_and_zoom?left='+(ww.g[0][1]-0.002)+'&right='+(ww.g[0][1]+0.002)+'&bottom='+(ww.g[0][0]-0.001)+'&top='+(ww.g[0][0]+0.001)+'" target="_blank" rel="noopener">JOSM</a>';
    var hi=ww.id===w.id?' style="background:#FFF8DC"':'';
    return '<div class="pp-row"'+hi+'><span class="tg '+ww.c.toLowerCase()+'">'+ww.c+'</span> way '+idLnk+' &middot; '+iD+' &middot; '+jo+'</div>';
  }).join('');
  var ids=segs.map(function(p){return p.wd.id;}).join(',');
  var html='<div class="pp-h">'+escAttr(name)+'</div>'
    +'<div class="pp-bd">'+segs.length+' segment'+(segs.length>1?'s':'')+(bdStr?' &middot; '+bdStr:'')+' &middot; '+(w.h||'')+'</div>'
    +'<div class="pp-clicked"><b>You clicked:</b> way '+w.id+' (Class '+w.c+(w.o==='yes'?', oneway=yes':'')+')</div>'
    +'<div class="pp-list">'+rows+'</div>'
    +'<div class="pp-foot">'
    +'<button class="pp-btn" onclick="loadInJOSM(['+ids+'])">Open all in JOSM</button>'
    +'<a href="#" onclick="event.preventDefault();clearHighlight();map.closePopup();">Clear highlight</a>'
    +'</div>';
  L.popup({maxWidth:340,autoPan:true}).setLatLng(latlng).setContent(html).openOn(map);
}

// Build polylines + gap markers (deferred so the loading overlay paints first).
setTimeout(function(){
  D.ways.forEach(function(w){
    if(!w.g||w.g.length<2)return;
    var p=L.polyline(w.g,{color:CL[w.c]||'#999',weight:CW[w.c]||1.5,opacity:w.c==='C'?0.4:0.8});
    p.wd=w;ly[w.c].addLayer(p);aP.push(p);
    if(w.n){(streetIndex[w.n]=streetIndex[w.n]||[]).push(p);}
    wayIndex[w.id]=p;
    p.on('click',function(e){
      L.DomEvent.stopPropagation(e);
      if(this.wd.n)setHighlight('street',this.wd.n);
      openStreetPopup(this.wd,e.latlng);
    });
  });

  D.gaps.forEach(function(g){
    var c=L.circleMarker([g.lat,g.lon],{radius:7,color:'#9C27B0',fillColor:'#CE93D8',fillOpacity:0.8,weight:2});
    c.bindPopup('<b>Probable Node Disconnect</b><br>Street: <b>'+escAttr(g.street||'')+'</b><br>Gap: '+g.distance_m+' m<br>'
      +'Way <a href="https://www.openstreetmap.org/way/'+g.way1_id+'" target="_blank" rel="noopener">'+g.way1_id+'</a> &harr; '
      +'<a href="https://www.openstreetmap.org/way/'+g.way2_id+'" target="_blank" rel="noopener">'+g.way2_id+'</a><br>'
      +'<a href="http://localhost:8111/load_and_zoom?left='+(g.lon-0.002)+'&right='+(g.lon+0.002)+'&bottom='+(g.lat-0.001)+'&top='+(g.lat+0.001)+'" target="_blank" rel="noopener">Open in JOSM</a>');
    c.on('click',function(e){
      L.DomEvent.stopPropagation(e);
      setHighlight('gap',[g.way1_id,g.way2_id]);
    });
    gL.addLayer(c);
  });

  ly.AB.addTo(map);ly.A.addTo(map);ly.B.addTo(map);gL.addTo(map);
  uvs();
  document.getElementById('loading').style.display='none';
  readHash();
},50);

var hp=D.ways.map(function(w){
  if(!w.g||!w.g.length)return null;
  var m=w.g[Math.floor(w.g.length/2)];
  return[m[0],m[1],w.c==='AB'?1:w.c==='A'?0.8:w.c==='B'?0.5:0.2];
}).filter(Boolean);

function tgl(id,layer){
  document.getElementById(id).addEventListener('change',function(){
    if(this.checked)map.addLayer(layer);else map.removeLayer(layer);
    uvs();scheduleHashWrite();
  });
}
tgl('lab',ly.AB);tgl('la',ly.A);tgl('lb',ly.B);tgl('lc',ly.C);tgl('lgap',gL);

document.getElementById('lheat').addEventListener('change',function(){
  if(this.checked){
    if(!hL)hL=L.heatLayer(hp,{radius:20,blur:15,maxZoom:16,gradient:{0.2:'blue',0.4:'cyan',0.6:'lime',0.8:'yellow',1.0:'red'}});
    map.addLayer(hL);
  }else if(hL)map.removeLayer(hL);
  scheduleHashWrite();
});

// Basemap selector init.
(function initBasemap(){
  var bmEl=document.getElementById('bmsel'),wbEl=document.getElementById('waybackRel');
  var saved='carto-voyager';
  try{saved=localStorage.getItem('tigerBmKey')||'carto-voyager';}catch(e){}
  if(bmEl.querySelector('option[value="'+saved+'"]'))bmEl.value=saved;
  updateWaybackVis();
  setBasemap(bmEl.value);
  bmEl.addEventListener('change',function(){setBasemap(this.value);});
  wbEl.addEventListener('change',function(){
    try{localStorage.setItem('tigerWaybackRel',this.value);}catch(e){}
    if(bmEl.value==='esri-wayback')setBasemap('esri-wayback');
  });
})();

// Sidebar collapse toggle (desktop).
(function initSidebarToggle(){
  var sb=document.getElementById('sidebar'),btn=document.getElementById('sbToggle');
  function apply(collapsed){
    if(collapsed){sb.classList.add('collapsed');btn.innerHTML='&#9654;';btn.title='Expand sidebar';}
    else{sb.classList.remove('collapsed');btn.innerHTML='&#9664;';btn.title='Collapse sidebar';}
    setTimeout(function(){if(map&&map.invalidateSize)map.invalidateSize();uvs();},280);
  }
  var startCollapsed=false;
  try{startCollapsed=localStorage.getItem('tigerSidebarCollapsed')==='1';}catch(e){}
  apply(startCollapsed);
  btn.addEventListener('click',function(){
    var nowCollapsed=!sb.classList.contains('collapsed');
    apply(nowCollapsed);
    try{localStorage.setItem('tigerSidebarCollapsed',nowCollapsed?'1':'0');}catch(e){}
  });
})();

map.on('click',function(){if(hl.type){clearHighlight();map.closePopup();}});

function uvs(){
  var b=map.getBounds(),v={t:0,AB:0,A:0,B:0,C:0,gp:0,st:new Set()};
  aP.forEach(function(p){var w=p.wd;if(w.g.length>0&&b.contains(L.latLng(w.g[0][0],w.g[0][1]))){v.t++;v[w.c]++;if(w.n)v.st.add(w.n);}});
  gL.eachLayer(function(m){if(b.contains(m.getLatLng()))v.gp++;});
  document.getElementById('vs').innerHTML='<b>'+v.t+'</b> segments in view<br><b>'+v.st.size+'</b> named streets<br>'
    +'AB: <b>'+v.AB+'</b> | A: <b>'+v.A+'</b><br>B: <b>'+v.B+'</b> | C: <b>'+v.C+'</b><br>Node gaps: <b>'+v.gp+'</b>';
}
map.on('moveend',function(){uvs();scheduleHashWrite();});

var si={};
D.ways.forEach(function(w){if(w.n&&!si[w.n]&&w.g&&w.g.length)si[w.n]=w.g[Math.floor(w.g.length/2)];});

var srActiveIdx=-1;
function srUpdateActive(idx){
  var sr=document.getElementById('sr');
  var opts=sr.querySelectorAll('[role="option"]');
  if(!opts.length){srActiveIdx=-1;document.getElementById('search').removeAttribute('aria-activedescendant');return;}
  if(idx<0)idx=opts.length-1;else if(idx>=opts.length)idx=0;
  srActiveIdx=idx;
  opts.forEach(function(el,i){
    var on=i===idx;
    el.setAttribute('aria-selected',on?'true':'false');
    if(on){el.scrollIntoView({block:'nearest'});}
  });
  document.getElementById('search').setAttribute('aria-activedescendant',opts[idx].id);
}
document.getElementById('search').addEventListener('input',function(){
  var q=this.value.toLowerCase(),r=document.getElementById('sr'),self=this;
  srActiveIdx=-1;self.removeAttribute('aria-activedescendant');
  if(q.length<2){r.innerHTML='';self.setAttribute('aria-expanded','false');return;}
  var m=Object.keys(si).filter(function(n){return n.toLowerCase().indexOf(q)>=0;}).slice(0,10);
  self.setAttribute('aria-expanded',m.length>0?'true':'false');
  r.innerHTML=m.map(function(n,i){
    return '<div role="option" tabindex="0" id="opt'+i+'" aria-selected="false" style="padding:3px 0;cursor:pointer;color:var(--link);font-size:11px" data-street="'+escAttr(n)+'">'+n+'</div>';
  }).join('');
  r.querySelectorAll('div').forEach(function(el){
    el.addEventListener('click',function(){flyTo(this.getAttribute('data-street'));});
    el.addEventListener('keydown',function(e){if(e.key==='Enter'||e.key===' '){e.preventDefault();flyTo(this.getAttribute('data-street'));}});
  });
});
document.getElementById('search').addEventListener('keydown',function(e){
  if(e.key==='ArrowDown'){e.preventDefault();srUpdateActive(srActiveIdx+1);}
  else if(e.key==='ArrowUp'){e.preventDefault();srUpdateActive(srActiveIdx-1);}
  else if(e.key==='Enter'){
    var sr=document.getElementById('sr');
    var opts=sr.querySelectorAll('[role="option"]');
    if(srActiveIdx>=0&&opts[srActiveIdx]){e.preventDefault();flyTo(opts[srActiveIdx].getAttribute('data-street'));}
  }
});

function flyTo(name){
  var c=si[name];if(c)map.setView(c,17);
  var s=document.getElementById('search');s.value=name;
  document.getElementById('sr').innerHTML='';
  s.setAttribute('aria-expanded','false');
  setHighlight('street',name);
}

function xGeo(){
  var b=map.getBounds(),f=[];
  aP.forEach(function(p){var w=p.wd;
    if(w.g.length>0&&b.contains(L.latLng(w.g[0][0],w.g[0][1]))){
      f.push({type:"Feature",properties:{id:w.id,name:w.n,highway:w.h,oneway:w.o,defect_class:w.c,severity:w.s},
        geometry:{type:"LineString",coordinates:w.g.map(function(pt){return[pt[1],pt[0]];})}});
    }
  });
  dl('tiger_export.geojson',JSON.stringify({type:"FeatureCollection",features:f},null,2),'application/geo+json');
}

function xCSV(){
  var b=map.getBounds(),r=['way_id,street,highway,oneway,defect_class,severity,osm_link'];
  aP.forEach(function(p){var w=p.wd;
    if(w.g.length>0&&b.contains(L.latLng(w.g[0][0],w.g[0][1]))){
      r.push([w.id,'"'+(w.n||'').replace(/"/g,'""')+'"',w.h||'',w.o||'',w.c,w.s,'https://www.openstreetmap.org/way/'+w.id].join(','));
    }
  });
  dl('tiger_export.csv',r.join(String.fromCharCode(10)),'text/csv');
}

function dl(n,c,t){var b=new Blob([c],{type:t}),a=document.createElement('a');a.href=URL.createObjectURL(b);a.download=n;a.click();}

// === R4 hamburger (mobile) ===
(function(){
  var hb=document.getElementById('hamburger'),bd=document.getElementById('backdrop'),sb=document.getElementById('sidebar');
  hb.addEventListener('click',function(){
    var open=sb.classList.toggle('open');
    bd.hidden=!open;hb.setAttribute('aria-expanded',open?'true':'false');
  });
  bd.addEventListener('click',function(){
    sb.classList.remove('open');bd.hidden=true;hb.setAttribute('aria-expanded','false');
  });
})();

// === E2 keyboard shortcuts + help overlay ===
function toggleLayer(id){document.getElementById(id).click();}
function toggleHelp(){helpOpen=!helpOpen;document.getElementById('help').hidden=!helpOpen;}

document.addEventListener('keydown',function(e){
  var t=e.target,inInput=t.tagName==='INPUT'||t.tagName==='TEXTAREA'||t.tagName==='SELECT';
  if(inInput&&e.key!=='Escape')return;
  if(e.key==='/'||(e.ctrlKey&&e.key.toLowerCase()==='k')){
    e.preventDefault();document.getElementById('search').focus();
  }else if(e.key==='Escape'){
    map.closePopup();clearHighlight();
    var s=document.getElementById('search');s.blur();s.value='';
    document.getElementById('sr').innerHTML='';
    s.setAttribute('aria-expanded','false');
    s.removeAttribute('aria-activedescendant');
    srActiveIdx=-1;
    if(helpOpen)toggleHelp();
  }else if(['1','2','3','4','5','6'].indexOf(e.key)>=0){
    var ids=['lab','la','lb','lc','lgap','lheat'];
    toggleLayer(ids[parseInt(e.key,10)-1]);
  }else if(e.key.toLowerCase()==='s'){
    // Toggle between current basemap and esri-imagery
    var bm=document.getElementById('bmsel');
    if(bm.value==='esri-imagery'){bm.value=prevBMKey||'carto-voyager';}
    else{prevBMKey=bm.value;bm.value='esri-imagery';}
    setBasemap(bm.value);
  }else if(e.key==='?'){
    toggleHelp();
  }
});

// === E1 URL hash state ===
function writeHash(){
  if(restoring)return;
  var c=map.getCenter(),z=map.getZoom();
  var on=['lab','la','lb','lc','lgap','lheat'].filter(function(id){return document.getElementById(id).checked;}).map(function(id){return LK[id];});
  var parts=['z='+z,'c='+c.lat.toFixed(5)+','+c.lng.toFixed(5),'l='+on.join(',')];
  if(curBMKey&&curBMKey!=='carto-voyager')parts.push('bm='+encodeURIComponent(curBMKey));
  if(document.documentElement.classList.contains('dark'))parts.push('dark=1');
  if(hl.type==='street')parts.push('hi=street:'+encodeURIComponent(hl.key).replace(/%20/g,'+'));
  try{history.replaceState(null,'','#'+parts.join('&'));}catch(e){}
}
function scheduleHashWrite(){clearTimeout(hashTimer);hashTimer=setTimeout(writeHash,200);}

function readHash(){
  if(!location.hash||location.hash.length<2)return;
  restoring=true;
  try{
    var p=new URLSearchParams(location.hash.slice(1));
    if(p.has('l')){
      var want=p.get('l').split(',');
      Object.keys(LR).forEach(function(k){
        var cb=document.getElementById(LR[k]);
        var should=want.indexOf(k)>=0;
        if(cb.checked!==should)cb.click();
      });
    }
    if(p.has('bm')){
      var bmKey=decodeURIComponent(p.get('bm'));
      var bmEl=document.getElementById('bmsel');
      if(bmEl.querySelector('option[value="'+bmKey+'"]')){bmEl.value=bmKey;setBasemap(bmKey);}
    }
    if(p.get('dark')==='1'&&!document.documentElement.classList.contains('dark'))setDark(true);
    if(p.has('c')&&p.has('z')){
      var cc=p.get('c').split(',').map(parseFloat);
      map.setView([cc[0],cc[1]],parseFloat(p.get('z')));
    }
    if(p.has('hi')){
      var hi=p.get('hi');
      if(hi.indexOf('street:')===0){
        var name=decodeURIComponent(hi.slice(7).replace(/\+/g,' '));
        if(streetIndex[name])setHighlight('street',name);
      }
    }
  }catch(err){console.warn('hash parse failed',err);}
  restoring=false;
}

// === E3 JOSM batch remote control ===
function showStatus(msg){var s=document.getElementById('status');s.textContent=msg;s.hidden=false;clearTimeout(s._t);s._t=setTimeout(function(){s.hidden=true;},4000);}

function probeJOSM(){
  if(josmProbeOk!==null)return Promise.resolve(josmProbeOk);
  if(location.protocol==='https:'){josmProbeOk=false;return Promise.resolve(false);}
  var opts={mode:'no-cors'};
  if(typeof AbortSignal!=='undefined'&&AbortSignal.timeout){try{opts.signal=AbortSignal.timeout(800);}catch(e){}}
  return fetch('http://localhost:8111/version',opts).then(function(){josmProbeOk=true;return true;}).catch(function(){josmProbeOk=false;return false;});
}

function loadInJOSM(wayIds){
  if(!wayIds||!wayIds.length){
    var b=map.getBounds();wayIds=[];
    aP.forEach(function(p){
      var w=p.wd;
      if(!map.hasLayer(ly[w.c]))return;
      if(w.g.length&&b.contains(L.latLng(w.g[0][0],w.g[0][1])))wayIds.push(w.id);
    });
  }
  if(!wayIds.length){showStatus('No visible ways to load');return;}
  probeJOSM().then(function(ok){
    if(!ok){
      showStatus(location.protocol==='https:'
        ?'JOSM remote control needs http:// or file:// (this page is loaded over https://).'
        :'JOSM not running \\u2014 start JOSM and enable Remote Control in Preferences \\u2192 Remote Control.');
      return;
    }
    var batches=[],cur=[],curLen=80;
    wayIds.forEach(function(id){
      var s='w'+id;
      // 7500 chars leaves headroom under the typical 8000-char URL limit.
      if(curLen+s.length+1>7500){batches.push(cur);cur=[];curLen=80;}
      cur.push(s);curLen+=s.length+1;
    });
    if(cur.length)batches.push(cur);
    function next(i){
      if(i>=batches.length){showStatus('Sent '+wayIds.length+' way'+(wayIds.length>1?'s':'')+' to JOSM ('+batches.length+' batch'+(batches.length>1?'es':'')+')');return;}
      fetch('http://localhost:8111/load_object?objects='+batches[i].join(',')+'&relation_members=true',{mode:'no-cors'})
        .finally(function(){setTimeout(function(){next(i+1);},500);});
    }
    next(0);
  });
}

// === E5 print mode ===
function printView(){
  var b=map.getBounds(),v={AB:0,A:0,B:0,C:0,gp:0};
  aP.forEach(function(p){var w=p.wd;if(w.g.length&&b.contains(L.latLng(w.g[0][0],w.g[0][1])))v[w.c]++;});
  gL.eachLayer(function(m){if(b.contains(m.getLatLng()))v.gp++;});
  document.getElementById('printBar').innerHTML='TIGER Audit \\u2014 '+ZONE_NAME+' \\u2014 '+new Date().toLocaleDateString()
    +' &middot; Visible: '+v.AB+' AB &middot; '+v.A+' A &middot; '+v.B+' B &middot; '+v.C+' C &middot; '+v.gp+' gaps';
  // Force light theme for printing so paper output stays readable, restore after.
  var wasDark=document.documentElement.classList.contains('dark');
  if(wasDark)document.documentElement.classList.remove('dark');
  document.body.classList.add('printing');
  setTimeout(function(){window.print();},250);
  window._tigerPrintWasDark=wasDark;
}
window.addEventListener('afterprint',function(){
  document.body.classList.remove('printing');
  if(window._tigerPrintWasDark){
    document.documentElement.classList.add('dark');
    window._tigerPrintWasDark=false;
  }
});

// === E6 dark mode (chrome theme; tiles are managed by basemap selector) ===
function setDark(on){
  document.documentElement.classList.toggle('dark',!!on);
  var btn=document.getElementById('dmt');
  btn.textContent=on?'\\u2600\\uFE0F':'\\uD83C\\uDF19';
  var lbl=on?'Switch to light mode':'Switch to dark mode';
  btn.setAttribute('aria-label',lbl);
  btn.setAttribute('title',lbl);
  try{localStorage.setItem('tiger-dark-mode',on?'1':'0');}catch(e){}
  scheduleHashWrite();
}
document.getElementById('dmt').addEventListener('click',function(){
  setDark(!document.documentElement.classList.contains('dark'));
});
(function(){
  var stored=null;try{stored=localStorage.getItem('tiger-dark-mode');}catch(e){}
  var prefersDark=window.matchMedia&&window.matchMedia('(prefers-color-scheme: dark)').matches;
  if(stored==='1'||(stored===null&&prefersDark))setDark(true);
})();
</script>
</body>
</html>
"""



def _bbox_center(bbox: tuple[float, float, float, float]) -> tuple[float, float]:
    s, w, n, e = bbox
    return ((s + n) / 2.0, (w + e) / 2.0)


def _round_geom(geom: list[list[float]]) -> list[list[float]]:
    # 7 decimals ~= 1 cm precision; reduces dashboard file size meaningfully.
    return [[round(p[0], 7), round(p[1], 7)] for p in geom]


def write_dashboard(classified: dict, zone_key: str, out_path: Path, audit_ts: str) -> None:
    zone = ZONES[zone_key]
    s = classified["summary_stats"]
    by_class = s.get("by_class", {})
    compact_stats = {
        "total": s["total"],
        "residential": s["residential"],
        "oneway": s["oneway_yes_total"],
        "class_a": by_class.get(CLASS_A, 0),
        "class_b": by_class.get(CLASS_B, 0),
        "class_ab": by_class.get(CLASS_AB, 0),
        "class_c": by_class.get(CLASS_C, 0),
        "multi_seg_streets": s["class_b_street_count"],
        "gaps_found": s.get("gaps_found", 0),
    }
    ways_compact = []
    for w in classified["all_ways"]:
        ways_compact.append({
            "id": w["id"],
            "h": w["highway"] or "",
            "n": w["name"] or "",
            "o": w["oneway"] or "",
            "c": w["defect_class"],
            "s": w["severity"],
            "g": _round_geom(w["geometry"]),
        })
    payload = {
        "zone_name": zone["name"],
        "audit_ts": audit_ts,
        "stats": compact_stats,
        "ways": ways_compact,
        "gaps": classified.get("gaps", []),
    }
    center = _bbox_center(zone["bbox"])
    html = (
        DASHBOARD_TEMPLATE
        .replace("__ZONE_NAME__", zone["name"])
        .replace("__AUDIT_TS__", audit_ts)
        .replace("__TOTAL__", f"{s['total']:,}")
        .replace("__CENTER__", f"[{center[0]},{center[1]}]")
        .replace("__DATA_JSON__", json.dumps(payload, ensure_ascii=False))
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")
    print(f"  Dashboard saved: {out_path}")


# ---------------------------------------------------------------------------
# Phase 5 - CSVs, README, console summary
# ---------------------------------------------------------------------------

CSV_FIELDS_ALL = [
    "id", "name", "highway", "oneway", "defect_class", "severity",
    "tiger_reviewed", "tiger_cfcc", "tiger_name_base", "surface", "lanes", "maxspeed",
]


def _write_csv(path: Path, rows: list[dict], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_csvs(classified: dict, csv_dir: Path) -> None:
    class_rank = {c: i for i, c in enumerate(CLASS_ORDER)}
    sorted_ways = sorted(
        classified["all_ways"],
        key=lambda w: (class_rank[w["defect_class"]], (w["name_display"] or "").lower(), w["id"] or 0),
    )
    _write_csv(csv_dir / "all_ways.csv", sorted_ways, CSV_FIELDS_ALL)
    _write_csv(
        csv_dir / "class_a_false_oneway.csv",
        sorted(classified["class_a"], key=lambda w: ((w["name_display"] or "").lower(), w["id"] or 0)),
        CSV_FIELDS_ALL,
    )

    # Per-street gap counts so volunteers can prioritize Class B streets
    # by the number of probable disconnects rather than just segment count.
    gaps_per_street: dict[str, int] = defaultdict(int)
    for g in classified.get("gaps", []):
        street = g.get("street")
        if street:
            gaps_per_street[street] += 1

    class_b_rows: list[dict] = []
    for street, ways in classified["class_b_streets"].items():
        gap_count = gaps_per_street.get(street, 0)
        for w in ways:
            class_b_rows.append({
                **w,
                "_street_segments": len(ways),
                "_street_gap_count": gap_count,
            })
    class_b_rows.sort(
        key=lambda w: (
            -w["_street_gap_count"],
            -w["_street_segments"],
            (w["name_display"] or "").lower(),
            w["id"] or 0,
        )
    )
    _write_csv(
        csv_dir / "class_b_multi_segment.csv",
        class_b_rows,
        CSV_FIELDS_ALL + ["_street_segments", "_street_gap_count"],
    )
    _write_csv(
        csv_dir / "class_ab_compound.csv",
        sorted(classified["class_ab"], key=lambda w: ((w["name_display"] or "").lower(), w["id"] or 0)),
        CSV_FIELDS_ALL,
    )
    print(f"  CSVs saved to {csv_dir}/")


def write_readme(out_dir: Path, zone_key: str, audit_ts: str) -> None:
    zone = ZONES[zone_key]
    text = (
        f"# TIGER/Line Audit - {zone['name']}\n\n"
        f"Read-only audit of OpenStreetMap ways with `tiger:reviewed=no` inside\n"
        f"the {zone['name']} MetroNow zone (bbox {zone['bbox']}).\n\n"
        f"Audit timestamp (UTC): {audit_ts}\n\n"
        f"## Re-run\n\n"
        f"```\n"
        f"python3 tiger_audit.py --zone {zone_key}\n"
        f"```\n\n"
        f"## Files\n\n"
        f"- `data/{zone_key}_raw_<UTC>.json` - raw Overpass response (preserved)\n"
        f"- `reports/TIGER-Audit-*.xlsx` - styled multi-sheet workbook\n"
        f"- `reports/TIGER-Audit-*-Dashboard.html` - interactive dashboard (open in browser)\n"
        f"- `csv/all_ways.csv` - master inventory\n"
        f"- `csv/class_a_false_oneway.csv` - residential + oneway=yes\n"
        f"- `csv/class_b_multi_segment.csv` - streets with 2+ unreviewed segments\n"
        f"- `csv/class_ab_compound.csv` - intersection of A and B\n"
    )
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "README.md").write_text(text, encoding="utf-8")


def print_summary(classified: dict, zone_key: str, audit_ts: str, out_dir: Path, xlsx_name: str, html_name: str, raw_filename: str) -> None:
    zone = ZONES[zone_key]
    s = classified["summary_stats"]
    p1_h = s["class_ab_count"] * 12 / 60
    p2_h = s["class_a_only_count"] * 4 / 60
    p3_h = s["class_b_street_count"] * 7 / 60
    p4_items = max(s["residential"] - s["class_a_count"], 0)
    p4_h = p4_items * 2 / 60
    p0p1 = p1_h
    total_h = p1_h + p2_h + p3_h + p4_h
    line = "=" * 44
    print()
    print(line)
    print(f"  TIGER Audit Complete: {zone['name']}")
    print(f"  Date: {audit_ts}")
    print(f"  Source: Overpass API (live query)")
    print(line)
    print(f"  Total unreviewed segments:    {s['total']:,}")
    print(f"  Residential (unreviewed):     {s['residential']:,}")
    print(f"  Class A (false one-way):      {s['class_a_count']:,}")
    print(f"  Class B (multi-segment):      {s['class_b_way_count']:,}")
    print(f"  Class AB (compound):          {s['class_ab_count']:,}")
    print(f"  Node gaps detected:           {s.get('gaps_found', 0):,}")
    print("-" * 44)
    print(f"  Est. volunteer hours (P0+P1): {p0p1:.1f}")
    print(f"  Est. volunteer hours (all):   {total_h:.1f}")
    print("-" * 44)
    print(f"  Files saved to: {out_dir}/")
    print(f"  XLSX:      reports/{xlsx_name}")
    print(f"  Dashboard: reports/{html_name}")
    print(f"  CSVs:      csv/ (4 files)")
    print(f"  Raw JSON:  data/{raw_filename}")
    print(line)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def _zone_paths(zone_key: str) -> tuple[Path, Path, Path, Path]:
    out_dir = SCRIPT_DIR / f"tiger_audit_{zone_key}"
    return (
        out_dir,
        out_dir / "reports",
        out_dir / "csv",
        out_dir / "data",
    )


def _zone_name_to_filename(zone_name: str) -> str:
    # Hyphens, not underscores — chat apps (Slack/Discord/Markdown) italicize
    # text between underscores, which mangles shared file URLs.
    return zone_name.replace(" / ", "-").replace(" ", "-").replace(".", "")


def run_zone(zone_key: str) -> dict:
    zone = ZONES[zone_key]
    out_dir, reports_dir, csv_dir, data_dir = _zone_paths(zone_key)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n=== {zone['name']} ({zone_key}) ===")
    print("Phase 1: Fetching from Overpass API...")
    raw = fetch_overpass(zone_key, out_dir)

    raw_files = sorted(data_dir.glob(f"{zone_key}_raw_*.json"))
    raw_name = raw_files[-1].name if raw_files else "(none)"

    print("Phase 2: Classifying defects...")
    classified = classify(raw)
    s = classified["summary_stats"]
    print(
        f"  total={s['total']} class_a={s['class_a_count']} "
        f"class_b_streets={s['class_b_street_count']} "
        f"class_b_ways={s['class_b_way_count']} class_ab={s['class_ab_count']}"
    )

    safe_name = _zone_name_to_filename(zone["name"])
    xlsx_name = f"TIGER-Audit-{safe_name}.xlsx"
    html_name = f"TIGER-Audit-{safe_name}-Dashboard.html"

    audit_ts = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")

    print("Phase 3: Writing XLSX...")
    write_xlsx(classified, zone_key, reports_dir / xlsx_name, overpass_query(zone["bbox"]), audit_ts)

    print("Phase 4: Writing HTML dashboard...")
    write_dashboard(classified, zone_key, reports_dir / html_name, audit_ts)

    print("Phase 5: Writing CSVs and README...")
    write_csvs(classified, csv_dir)
    write_readme(out_dir, zone_key, audit_ts)

    print_summary(classified, zone_key, audit_ts, out_dir, xlsx_name, html_name, raw_name)
    return {
        "zone_key": zone_key,
        "zone_name": zone["name"],
        "out_dir": out_dir,
        "classified": classified,
        "audit_ts": audit_ts,
    }


def write_combined_dashboard(results: list[dict], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    # Zone bboxes overlap (Northgate ↔ Forest Park ~32 km², Springdale ↔
    # Forest Park ~14 km²). The same OSM way appears in multiple per-zone
    # audits when it sits in an overlap region, so we dedupe by way ID
    # before computing combined stats — otherwise the totals double-count.
    seen_way_ids: set[int] = set()
    merged_ways: list[dict] = []
    for r in results:
        c = r["classified"]
        for w in c["all_ways"]:
            wid = w["id"]
            if wid is None or wid in seen_way_ids:
                continue
            seen_way_ids.add(wid)
            merged_ways.append(w)

    # Cross-zone class_b_streets: count unique physical streets, where two
    # same-named groups in different zones are treated as the same physical
    # street only if they share at least one OSM way ID. This is a
    # connected-components count over (street_name, way_id_set) groups — so
    # an arterial that crosses a zone boundary collapses to one street, while
    # two unrelated streets that happen to share a name (e.g. "Park Ave"
    # appearing in both Northgate and Forest Park with disjoint way IDs)
    # are counted separately.
    streets_by_name: dict[str, list[set[int]]] = defaultdict(list)
    for r in results:
        for street, ways in r["classified"]["class_b_streets"].items():
            wids = {w["id"] for w in ways if w["id"] is not None}
            if wids:
                streets_by_name[street].append(wids)
    class_b_streets_total = 0
    for _name, groups in streets_by_name.items():
        components: list[set[int]] = []
        for grp in groups:
            # A group can bridge multiple existing components when it shares
            # way IDs with two or more of them. Absorb every intersecting
            # component into a single merged set, then reinsert.
            merged = set(grp)
            survivors: list[set[int]] = []
            for comp in components:
                if comp & merged:
                    merged |= comp
                else:
                    survivors.append(comp)
            survivors.append(merged)
            components = survivors
        class_b_streets_total += len(components)

    # Same problem with gaps — a node disconnect inside an overlap region
    # is detected once per audit. First pass: dedupe by the unordered
    # way-pair key (catches identical pairs across zones).
    seen_gap_pairs: set[tuple[int, int]] = set()
    pair_deduped_gaps: list[dict] = []
    for r in results:
        for g in r["classified"].get("gaps", []):
            try:
                key = tuple(sorted([int(g["way1_id"]), int(g["way2_id"])]))
            except (KeyError, TypeError, ValueError):
                pair_deduped_gaps.append(g)
                continue
            if key in seen_gap_pairs:
                continue
            seen_gap_pairs.add(key)
            pair_deduped_gaps.append(g)

    # Second pass: cross-zone spatial clustering. A junction in a zone-
    # bbox overlap region may produce DIFFERENT representative way-pairs
    # in each zone (zone A picks A↔B, zone B picks B↔C for the same
    # physical missing node). The way-pair dedup above misses those, so
    # collapse remaining gaps within GAP_CLUSTER_M that share a street.
    pair_deduped_gaps.sort(key=lambda x: x.get("distance_m", 0))
    merged_gaps: list[dict] = []
    for g in pair_deduped_gaps:
        collapsed = False
        for fg in merged_gaps:
            if fg.get("street") != g.get("street"):
                continue
            if (
                _haversine_m(g["lat"], g["lon"], fg["lat"], fg["lon"])
                <= GAP_CLUSTER_M
            ):
                collapsed = True
                break
        if not collapsed:
            merged_gaps.append(g)

    # Stats are computed from the deduped list, not summed across zones.
    total = len(merged_ways)
    residential = sum(1 for w in merged_ways if w["highway"] == "residential")
    oneway_yes = sum(1 for w in merged_ways if w["oneway"] == "yes")
    by_class: dict[str, int] = defaultdict(int)
    for w in merged_ways:
        by_class[w["defect_class"]] += 1

    naive_total = sum(r["classified"]["summary_stats"]["total"] for r in results)
    duplicate_count = naive_total - total
    print(
        f"  Combined dashboard dedup: {naive_total:,} per-zone-sum -> "
        f"{total:,} unique ({duplicate_count:,} ways in overlapping bboxes)"
    )

    audit_ts = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")
    # NOTE: compact_stats semantics differ from per-zone summary_stats and the
    # all_zones_summary.csv columns. Here `class_a` / `class_b` / `class_ab` /
    # `class_c` are DISJOINT counts (a way is counted in exactly one), to feed
    # the dashboard's separate KPI cards. The CSV's `class_a_count` column
    # follows the per-zone convention where it includes AB ways (matching
    # `summary_stats["class_a_count"] = len(class_a)` from `classify()`). Both
    # are correct for their artifact; do not "reconcile" them by changing the
    # dashboard or the CSV — they intentionally measure different things.
    compact_stats = {
        "total": total,
        "residential": residential,
        "oneway": oneway_yes,
        "class_a": by_class.get(CLASS_A, 0),
        "class_b": by_class.get(CLASS_B, 0),
        "class_ab": by_class.get(CLASS_AB, 0),
        "class_c": by_class.get(CLASS_C, 0),
        "multi_seg_streets": class_b_streets_total,
        "gaps_found": len(merged_gaps),
    }
    ways_compact = [
        {
            "id": w["id"],
            "h": w["highway"] or "",
            "n": w["name"] or "",
            "o": w["oneway"] or "",
            "c": w["defect_class"],
            "s": w["severity"],
            "g": _round_geom(w["geometry"]),
        }
        for w in merged_ways
    ]
    payload = {
        "zone_name": "All Zones",
        "audit_ts": audit_ts,
        "stats": compact_stats,
        "ways": ways_compact,
        "gaps": merged_gaps,
    }
    # Center on the union of zone bboxes
    south = min(ZONES[r["zone_key"]]["bbox"][0] for r in results)
    west = min(ZONES[r["zone_key"]]["bbox"][1] for r in results)
    north = max(ZONES[r["zone_key"]]["bbox"][2] for r in results)
    east = max(ZONES[r["zone_key"]]["bbox"][3] for r in results)
    center = ((south + north) / 2.0, (west + east) / 2.0)

    html = (
        DASHBOARD_TEMPLATE
        .replace("__ZONE_NAME__", "All Zones (combined)")
        .replace("__AUDIT_TS__", audit_ts)
        .replace("__TOTAL__", f"{total:,}")
        .replace("__CENTER__", f"[{center[0]},{center[1]}]")
        .replace("__DATA_JSON__", json.dumps(payload, ensure_ascii=False))
    )
    out_path = out_dir / "TIGER-Audit-All-Zones-Dashboard.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"  Combined dashboard: {out_path}")

    csv_path = out_dir / "all_zones_summary.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow([
            "zone", "zone_key", "total", "residential",
            "class_a_count", "class_b_street_count", "class_b_way_count",
            "class_ab_count", "gaps_found",
        ])
        for r in results:
            s = r["classified"]["summary_stats"]
            w.writerow([
                r["zone_name"], r["zone_key"], s["total"], s["residential"],
                s["class_a_count"], s["class_b_street_count"], s["class_b_way_count"],
                s["class_ab_count"], s.get("gaps_found", 0),
            ])
        # Per-zone-summed totals (with overlap duplicates) AND deduped unique
        # totals on the bottom two rows, so the overlap is visible in the data.
        sum_total = sum(r["classified"]["summary_stats"]["total"] for r in results)
        sum_resi = sum(r["classified"]["summary_stats"]["residential"] for r in results)
        sum_a = sum(r["classified"]["summary_stats"]["class_a_count"] for r in results)
        sum_b_streets = sum(r["classified"]["summary_stats"]["class_b_street_count"] for r in results)
        sum_b_ways = sum(r["classified"]["summary_stats"]["class_b_way_count"] for r in results)
        sum_ab = sum(r["classified"]["summary_stats"]["class_ab_count"] for r in results)
        sum_gaps = sum(r["classified"]["summary_stats"].get("gaps_found", 0) for r in results)
        class_b_way_unique = sum(
            1 for ww in merged_ways
            if ww["defect_class"] in (CLASS_B, CLASS_AB)
        )
        w.writerow([
            "Hamilton County (per-zone sum, includes overlap duplicates)", "_sum",
            sum_total, sum_resi, sum_a, sum_b_streets, sum_b_ways, sum_ab, sum_gaps,
        ])
        w.writerow([
            "Hamilton County (unique, deduped by way ID)", "_unique",
            total, residential,
            by_class.get(CLASS_A, 0) + by_class.get(CLASS_AB, 0),
            class_b_streets_total, class_b_way_unique,
            by_class.get(CLASS_AB, 0), len(merged_gaps),
        ])
    print(f"  Combined summary CSV: {csv_path}")


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="TIGER/Line audit for SORTA MetroNow zones")
    parser.add_argument(
        "--zone",
        choices=list(ZONES.keys()) + ["all"],
        default="blue_ash_montgomery",
    )
    args = parser.parse_args(argv)

    if args.zone == "all":
        results = [run_zone(k) for k in ZONES]
        write_combined_dashboard(results, SCRIPT_DIR / "tiger_audit_all_zones")
    else:
        run_zone(args.zone)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
